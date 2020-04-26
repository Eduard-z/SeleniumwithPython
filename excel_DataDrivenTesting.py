import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)


def readData(file, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columno).value


def writeData (file, sheetName, rownum, columno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(file)


from selenium import webdriver
import time

driver = webdriver.Chrome(executable_path="C:/Users\zakharove\drivers\chromedriver.exe")
driver.get("https://login.salesforce.com/")
driver.maximize_window()

path = r"C:\Users\zakharove\Desktop\test.xlsx"

rows = getRowCount(path, 'Sheet3')
for r in range(2, rows+1):
    username = readData(path, 'Sheet3', r, 1)
    password = readData(path, 'Sheet3', r, 2)

    driver.find_element_by_name("username").send_keys(username)
    driver.find_element_by_name("pw").send_keys(password)
    driver.find_element_by_name("Login").click()

    if driver.title == "Salesforce":
        print("test passed")
        writeData(path, 'Sheet3', r, 3, "test passed")
    else:
        print("test failed")
        writeData(path, 'Sheet3', r, 3, "test failed")

    time.sleep(2)
    driver.get("https://login.salesforce.com/")
