import openpyxl

path = r"C:\Users\zakharove\Desktop\test.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("Sheet1")  # or sheet=workbook.active or sheet=workbook["Sheet1"]

# read data from MS Excel
rows = sheet.max_row  # 6
columns = sheet.max_column  # 9
print(rows, columns)

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r, column=c).value, end='    ')
    print()

# write data into MS Excel
sheet = workbook["Sheet2"]
for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(row=r, column=c).value = "welcome"

workbook.save(path)  # the file must not be opened while saving
